from openpyxl import load_workbook
import csv

# data preprocessor used to populate the excel sheet of zeroes with the raw data
def process_data(dataset, file_name) -> None:
    # access proper excel file for preprocessing
    if dataset == "train":
        text_workbook = load_workbook("data/text-train.xlsx")
        dataset_workbook = load_workbook("data/zeroes-train.xlsx")
    elif dataset == "test":
        text_workbook = load_workbook("data/text-test.xlsx")
        dataset_workbook = load_workbook("data/zeroes-test.xlsx")
    else:
        raise ValueError("dataset must be train or test")

    text = text_workbook.active
    dataset = dataset_workbook.active

    # excluded less common labels for now
    labels = {
        "None": 1,
        "Python and Coding": 2,
        "Github": 3,
        "MySQL": 4,
        "Assignments ": 5,  # space after "Assignments" is not a typo;
        # funnily enough, in the original dataset's "Issue" dropdown,
        # there is a stray space after the issues "Assignments", "Quizzes",
        # and "Understanding requirements and instructions". It is just
        # easier to fix it this way.
        "Quizzes ": 6,
        "Understanding requirements and instructions ": 7,
        "Learning New Material": 8,
        "Course Structure and Materials": 9,
        "Time Management and Motivation": 10,
        "Group Work": 11,
        "API": 12,
        "Project": 13
    }

    response_num = 2

    current_response = text.cell(row=2, column=2).value

    # iterate over zeroes and text spreadsheets concurrently and populate as needed
    # zeroes excel file will become the dataset we use
    reflections = [current_response]
    for row in text.iter_rows(min_row=2, min_col=1, max_row=text.max_row, max_col=2, values_only=True):
        if current_response != row[1]:
            current_response = row[1]
            reflections.append(current_response)
            response_num += 1
        try:
            dataset.cell(row=response_num, column=labels[row[0]]).value = 1
        except KeyError:
            print("Key not found, skipping label")
    print(len(reflections))

    # save populated dataset
    with open(file_name, "w", encoding="utf-8", newline="") as f:
        c = csv.writer(f)
        i = -1
        for r in dataset.rows:
            # skip label row
            if i == -1:
                i += 1
                continue
            l = []
            for cell in r:
                if cell.value is not None:
                    l.append(int(cell.value))
            l.append(reflections[i])
            c.writerow(l)
            i += 1





